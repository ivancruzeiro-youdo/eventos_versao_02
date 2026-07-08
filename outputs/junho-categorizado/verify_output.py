import openpyxl


path = "Fechamento JUNHO 2026 - categorizado.xlsx"
wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
ws = wb["Cópia de GASTOS DETALHADOS"]

data = 0
complete = 0
partial = 0
blank = 0

for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
    if all(value is None for value in row):
        continue
    data += 1
    space = row[4]
    sector = row[5]
    if space not in (None, "") and sector not in (None, ""):
        complete += 1
    elif space in (None, "") and sector in (None, ""):
        blank += 1
    else:
        partial += 1

print({"data_rows": data, "complete_space_sector": complete, "partial": partial, "blank_space_sector": blank})
for row in ws.iter_rows(min_row=2, max_row=12, max_col=7, values_only=True):
    print(row)
