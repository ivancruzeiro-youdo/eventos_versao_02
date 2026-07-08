import json
import re
import unicodedata
from collections import Counter, defaultdict

import openpyxl


SOURCE = "source.xlsx"
MAY_SHEET = "GASTOS DETALHADOS"
JUNE_SHEET = "Cópia de GASTOS DETALHADOS"
OUT = "updates.json"


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_pi(value):
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_empty_row(row):
    return all(cell.value is None for cell in row[:7])


def category_pair(row):
    space = row[4].value
    sector = row[5].value
    if space in (None, "") or sector in (None, ""):
        return None
    return (str(space).strip(), str(sector).strip())


wb = openpyxl.load_workbook(SOURCE, data_only=False)
may = wb[MAY_SHEET]
june = wb[JUNE_SHEET]

mapping = defaultdict(Counter)
may_data_rows = 0
for row in may.iter_rows(min_row=2, max_col=7):
    if is_empty_row(row):
        continue
    may_data_rows += 1
    pair = category_pair(row)
    if pair is None:
        continue
    key = (normalize_pi(row[0].value), normalize_text(row[3].value))
    if key[0] and key[1]:
        mapping[key][pair] += 1

safe_mapping = {}
ambiguous_keys = {}
for key, counts in mapping.items():
    if len(counts) == 1:
        safe_mapping[key] = next(iter(counts))
    else:
        ambiguous_keys["|".join(key)] = {
            f"{space}||{sector}": count for (space, sector), count in counts.items()
        }

updates = []
june_data_rows = 0
already_complete = 0
unmatched = []
conflicts = []

for row in june.iter_rows(min_row=2, max_col=7):
    if is_empty_row(row):
        continue
    june_data_rows += 1
    row_num = row[0].row
    existing_space = row[4].value
    existing_sector = row[5].value
    if existing_space not in (None, "") and existing_sector not in (None, ""):
        already_complete += 1
        continue
    key = (normalize_pi(row[0].value), normalize_text(row[3].value))
    category = safe_mapping.get(key)
    if not category:
        unmatched.append(
            {
                "row": row_num,
                "pi": row[0].value,
                "supplier": row[3].value,
                "nf": row[6].value,
            }
        )
        continue
    space, sector = category
    if existing_space not in (None, "") and str(existing_space).strip() != space:
        conflicts.append({"row": row_num, "field": "ESPAÇO", "existing": existing_space, "suggested": space})
        continue
    if existing_sector not in (None, "") and str(existing_sector).strip() != sector:
        conflicts.append({"row": row_num, "field": "SETOR", "existing": existing_sector, "suggested": sector})
        continue
    updates.append(
        {
            "row": row_num,
            "space": None if existing_space not in (None, "") else space,
            "sector": None if existing_sector not in (None, "") else sector,
            "pi": row[0].value,
            "supplier": row[3].value,
        }
    )

summary = {
    "may_data_rows": may_data_rows,
    "june_data_rows": june_data_rows,
    "safe_keys": len(safe_mapping),
    "ambiguous_keys": len(ambiguous_keys),
    "updates": len(updates),
    "already_complete": already_complete,
    "unmatched": len(unmatched),
    "conflicts": len(conflicts),
    "sample_updates": updates[:20],
    "sample_unmatched": unmatched[:50],
    "sample_conflicts": conflicts[:20],
}

with open(OUT, "w", encoding="utf-8") as f:
    json.dump({"summary": summary, "updates": updates, "ambiguous_keys": ambiguous_keys}, f, ensure_ascii=False, indent=2)

print(json.dumps(summary, ensure_ascii=False, indent=2))
