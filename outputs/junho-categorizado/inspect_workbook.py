import openpyxl


path = "source.xlsx"
wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
print(wb.sheetnames)

for sheet_name in ["GASTOS DETALHADOS", "Cópia de GASTOS DETALHADOS"]:
    ws = wb[sheet_name]
    print(f"SHEET {ws.title!r} rows={ws.max_row} cols={ws.max_column}")
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        print(row)
    print("---")
